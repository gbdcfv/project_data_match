from typing import Dict, Any, Union
from openpyxl import load_workbook
import openpyxl
from pyautocad import Autocad
import dxfgrabber
import os
import sys
from pathlib import Path
import pandas as pd

wb = openpyxl.Workbook()
ws = wb.active
acad = Autocad(create_if_not_exists=True)
acad.prompt("Hello! AutoCAD from pyautocad.")
print(acad.doc.Name)

file_name = 'process drawing.dxf'
acad.ActiveDocument.SaveAs(file_name, 25)
data = []
q = 0
dxf = dxfgrabber.readfile(file_name)

ent = dxf.entities
for i in dxf.entities:
    if i.dxftype == 'INSERT' and len(i.attribs) != 0:
        # pringt(([x.text for x in i.attribs])
        att = ""
        y = "-"
        dat: Dict[Union[str, Any], Union[str, Any]] = {'块名': i.name, '插入点': 'x：%d;y：%d;z：%d' % i.insert, y: ''}
        for atts in i.attribs:
            dat[atts.tag] = atts.text
        data.append(dat)

datt = []
row = 1
for a in data:
    row += 1
    for key, val in a.items():
        if key not in datt:
            datt += [key]
        col = datt.index(key) + 1
        ws.cell(row=row, column=col, value=val)
        for pp in range(len(datt)):
            ws.cell(row=1, column=pp + 1, value=datt[pp])
wb.save(Path(sys.argv[0]).parent.joinpath('统计表.xlsx'))

list_n = ['支架编号', '管线号', '计算号', '节点号', '摩擦系数']
df = pd.read_excel('统计表.xlsx', usecols=list_n)
df = df.dropna(axis=0, how='all')
df = df.reindex(columns=['支架编号', '区号', '管线号', '管线尺寸', '保冷厚度', '支架长度', '节点号', '计算号', '摩擦系数'])
df.replace(0.1, 'Y', inplace=True)
df.replace(0.3, 'N', inplace=True)
print(df)
df.to_excel('333.xlsx')
wb = load_workbook('333.xlsx')
ws = wb.active
rc = ws['b']
rr = ws['2']

file_name = input("请输入支撑汇总表文件名：")
wb1 = load_workbook('' + file_name + '.xlsx')
ws1 = wb1['support']
col_range1 = ws1['j']

for row in range(2, len(rr) + 1):
    for col in range(2, len(rc) + 1):
        dt = ws.cell(col, row).value
        ws1.cell(col + 1, row + 1).value = dt
wb1.save('' + file_name + '.xlsx')


for x in range(3, len(col_range1) + 1):
    ws1 = wb1['support']
    if ws1.cell(x, 11).value is None:
        x = x + 1
    else:
        nn = ws1.cell(x, 10).value
        mm = ws1.cell(x, 9).value
        wb2 = load_workbook('' + nn + '.xlsx')
        ws2 = wb2['Restraint Summary Extended']
        col_range2 = ws2['b']
        k = 0
        for case in range(len(col_range2) - 1, -1, -1):
            k = k + 1
            if ws2.cell(case, 2).value is None:
                break
        for y in range(1, len(col_range2) + 1):
            if ws2.cell(y, 1).value == mm:
                list_fx = []
                list_fy = []
                list_fz = []
                list_mx = []
                list_my = []
                list_mz = []
                list_dx = []
                list_dy = []
                list_dz = []
                for i in range(1, k):
                    list_fx.append(ws2.cell(y + i, 3).value)
                    list_fy.append(ws2.cell(y + i, 4).value)
                    list_fz.append(ws2.cell(y + i, 5).value)
                    list_mx.append(ws2.cell(y + i, 6).value)
                    list_my.append(ws2.cell(y + i, 7).value)
                    list_mz.append(ws2.cell(y + i, 8).value)
                    list_dx.append(ws2.cell(y + i, 9).value)
                    list_dy.append(ws2.cell(y + i, 10).value)
                    list_dz.append(ws2.cell(y + i, 11).value)
                c = ws2.cell(y, 3).value
                list_fx1 = []
                list_fy1 = []
                list_fz1 = []
                list_mx1 = []
                list_my1 = []
                list_mz1 = []
                list_dx1 = []
                list_dy1 = []
                list_dz1 = []
                for i in list_fx:
                    list_fx1.append(abs(i))
                n = max(list_fx1)
                o = list_fx[list_fx1.index(n)]
                for i in list_fy:
                    list_fy1.append(abs(i))
                n = max(list_fy1)
                p = list_fy[list_fy1.index(n)]
                for i in list_fz:
                    list_fz1.append(abs(i))
                n = max(list_fz1)
                q = list_fz[list_fz1.index(n)]
                for i in list_mx:
                    list_mx1.append(abs(i))
                n = max(list_mx1)
                r = list_mx[list_mx1.index(n)]
                for i in list_my:
                    list_my1.append(abs(i))
                n = max(list_my1)
                s = list_my[list_my1.index(n)]
                for i in list_mz:
                    list_mz1.append(abs(i))
                n = max(list_mz1)
                t = list_mz[list_mz1.index(n)]
                for i in list_dx:
                    list_dx1.append(abs(i))
                n = max(list_dx1)
                u = list_dx[list_dx1.index(n)]
                for i in list_dy:
                    list_dy1.append(abs(i))
                n = max(list_dy1)
                v = list_dy[list_dy1.index(n)]
                for i in list_dz:
                    list_dz1.append(abs(i))
                n = max(list_dz1)
                w = list_dz[list_dz1.index(n)]
                ws1 = wb1['support']
                ws1.cell(x, 12).value = c
                ws1.cell(x, 13).value = o
                # print('计算号:',nn,',节点:',mm,'FX最大值',o,'N已完成填写。')
                ws1.cell(x, 15).value = p
                # print('计算号:',nn,',节点:',mm,'FY最大值',p,'N已完成填写。')
                ws1.cell(x, 17).value = q
                # print('计算号:',nn,',节点:',mm,'FZ最大值',q,'N已完成填写。')
                ws1.cell(x, 19).value = r
                # print('计算号:',nn,',节点:',mm,'MX最大值',r,'N.M已完成填写。')
                ws1.cell(x, 20).value = s
                # print('计算号:',nn,',节点:',mm,'MY最大值',s,'N.M已完成填写。')
                ws1.cell(x, 21).value = t
                # print('计算号:',nn,',节点:',mm,'MZ最大值',t,'N.M已完成填写。')
                ws1.cell(x, 22).value = u
                # print('计算号:',nn,',节点:',mm,'DX最大值',u,'mm已完成填写。')
                ws1.cell(x, 23).value = v
                # print('计算号:',nn,',节点:',mm,'DY最大值',v,'mm已完成填写。')
                ws1.cell(x, 24).value = w
                # print('计算号:',nn,',节点:',mm,'DZ最大值',w,'mm已完成填写。')
        print('计算号:', nn, ',节点:', mm, '受力以及位移最大值信息已完成填写。')
wb1.save('' + file_name + '.xlsx')
os.remove(Path(sys.argv[0]).parent.joinpath('process drawing.dxf'))
os.remove(Path(sys.argv[0]).parent.joinpath('333.xlsx'))
os.remove(Path(sys.argv[0]).parent.joinpath('统计表.xlsx'))
print('已完成文件', file_name, '受力以及位移信息填写.')
end_src = input("请输入任意内容关闭程序：")
